"""
src/exporter/export_manager.py
==============================
Main exporter orchestrator. Gathers completed/failed datasets from the database,
triggers CSV/Excel writers, and initiates report compilation.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Any, List, Optional

from src.database.repository import BaseRepository
from .csv_exporter import CSVExporter
from .excel_exporter import ExcelExporter
from .report_generator import ReportGenerator
from .export_metrics import ExportMetrics

logger = logging.getLogger(__name__)


class ExportManager:
    """
    Orchestrates the complete dataset export process.
    """

    def __init__(
        self,
        repository: BaseRepository,
        metrics_file: str = "logs/export_metrics.json"
    ) -> None:
        self.repo = repository
        self.metrics = ExportMetrics(metrics_file)
        self.csv_exporter = CSVExporter(self.metrics)
        self.excel_exporter = ExcelExporter(self.metrics)

    def export_all(self, export_dir: str = "data/export", original_file_path: Optional[str] = None) -> Dict[str, str]:
        """
        Loads dataset records from the repository and runs the exports/reports.
        
        Returns:
            Dict containing the file paths generated.
        """
        logger.info(f"[ExportManager] Initiating complete export to directory: {export_dir}")
        export_path = Path(export_dir)
        export_path.mkdir(parents=True, exist_ok=True)

        generated_files = {}

        try:
            # 1. Fetch completed, failed, and retry records from database
            completed_profiles = self.repo.get_all_completed()
            failed_records = self.repo.get_all_failed()
            retry_records = self.repo.get_all_retries()
            
            # 2. Export completed contacts to CSV and Excel
            completed_csv_path = export_path / "completed_contacts.csv"
            self.csv_exporter.export(completed_profiles, str(completed_csv_path))
            generated_files["completed_csv"] = str(completed_csv_path)

            completed_xlsx_path = export_path / "completed_contacts.xlsx"
            self.excel_exporter.export(completed_profiles, str(completed_xlsx_path))
            generated_files["completed_excel"] = str(completed_xlsx_path)

            # 3. Export failed records list to CSV for raw tracking
            if failed_records:
                import csv
                failed_csv_path = export_path / "failed_records.csv"
                with open(failed_csv_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(["id", "npi", "company_name", "website", "error_message", "failed_at"])
                    for fr in failed_records:
                        writer.writerow([
                            fr.get("id"),
                            fr.get("npi") or "",
                            fr.get("company_name") or "",
                            fr.get("website") or "",
                            fr.get("error_message") or "",
                            fr.get("failed_at") or ""
                        ])
                generated_files["failed_csv"] = str(failed_csv_path)
                logger.info(f"[ExportManager] Exported {len(failed_records)} raw failed records to CSV.")

            # 4. Filter resolved duplicates from database (records marked Merged or having duplicate resolution)
            # Find Completed profiles where extraction_method == 'Merged' as database-level duplicates
            db_duplicates = []
            for p in completed_profiles:
                if p.extraction_method == "Merged":
                    db_duplicates.append({
                        "npi": getattr(p, "npi", "N/A"),
                        "company_name": p.business_name,
                        "website": p.official_website,
                        "resolution": f"Identified as duplicate (website/NPI match). Merged details. Confidence: {p.confidence}"
                    })

            # 5. Generate Markdown reports
            summary_path = export_path / "summary_report.md"
            # Get pending retries count
            pending_retries_count = len([r for r in retry_records if r.get("status") == "pending"])
            
            ReportGenerator.generate_summary_report(
                completed_count=len(completed_profiles),
                failed_count=len(failed_records),
                retry_count=pending_retries_count,
                duplicate_count=len(db_duplicates),
                filepath=str(summary_path)
            )
            generated_files["summary_report"] = str(summary_path)

            failed_report_path = export_path / "failed_records_report.md"
            ReportGenerator.generate_failed_report(failed_records, str(failed_report_path))
            generated_files["failed_report"] = str(failed_report_path)

            duplicate_report_path = export_path / "duplicate_report.md"
            ReportGenerator.generate_duplicate_report(db_duplicates, str(duplicate_report_path))
            generated_files["duplicate_report"] = str(duplicate_report_path)

            stats_report_path = export_path / "statistics_report.md"
            # Trigger stats report gathering JSON files
            ReportGenerator.generate_statistics_report(
                filepath=str(stats_report_path)
            )
            generated_files["statistics_report"] = str(stats_report_path)

            if original_file_path:
                self.update_original_excel(original_file_path)

            logger.info(f"[ExportManager] Export process finalized successfully. Generated {len(generated_files)} files.")
        except Exception as e:
            logger.error(f"[ExportManager] Export execution failed: {e}")
            raise e

        return generated_files

    def update_original_excel(self, file_path: str) -> None:
        """
        Updates the original input Excel file in-place with enriched details,
        preserving all styles and sheets.
        """
        import openpyxl
        import os
        from datetime import datetime
        
        logger.info(f"[ExportManager] Updating original Excel file in-place: {file_path}")
        if not os.path.exists(file_path):
            logger.warning(f"Original Excel file not found for in-place update: {file_path}")
            return
            
        try:
            # 1. Fetch completed rows from DB directly using session
            session = self.repo.conn_mgr.get_session()
            try:
                from src.database.database_manager import CompletedContactModel
                db_rows = session.query(CompletedContactModel).all()
                # lookup: {npi_str: row}
                lookup = {}
                for r in db_rows:
                    if r.npi:
                        lookup[str(r.npi).strip()] = r
            finally:
                session.close()
                
            if not lookup:
                logger.info("[ExportManager] No completed profiles with NPI found in DB to update Excel.")
                return

            # 2. Open Workbook using openpyxl
            wb = openpyxl.load_workbook(file_path)
            
            # Select target sheet: prioritize 'Investor Contacts' or first sheet
            sheet_name = "Investor Contacts"
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
                
            ws = wb[sheet_name]
            
            # 3. Detect column indices (1-indexed in openpyxl)
            headers = [cell.value for cell in ws[1]]
            
            # Map column name to 1-based index
            col_map = {}
            for idx, h in enumerate(headers):
                if h:
                    col_map[str(h).strip().lower()] = idx + 1
                    
            # Check required columns
            npi_col = col_map.get("npi")
            if not npi_col:
                logger.error("[ExportManager] 'NPI' column not found in Excel sheet headers.")
                return
                
            # Column mappings in the sheet
            phone_col = col_map.get("phone")
            email_col = col_map.get("email")
            website_col = col_map.get("source website")
            conf_col = col_map.get("confidence")
            updated_col = col_map.get("updated")
            status_col = col_map.get("status")
            
            # 4. Iterate rows and update cells (starting at row 2)
            updated_count = 0
            for r_idx in range(2, ws.max_row + 1):
                npi_val = ws.cell(row=r_idx, column=npi_col).value
                if npi_val is None:
                    continue
                npi_str = str(npi_val).strip()
                
                if npi_str in lookup:
                    db_row = lookup[npi_str]
                    
                    # Update fields
                    emails_list = db_row.emails or []
                    phones_list = db_row.phones or []
                    
                    emails_str = ", ".join(emails_list) if isinstance(emails_list, list) else str(emails_list)
                    phones_str = ", ".join(phones_list) if isinstance(phones_list, list) else str(phones_list)
                    
                    if email_col:
                        ws.cell(row=r_idx, column=email_col).value = emails_str or None
                    if phone_col:
                        ws.cell(row=r_idx, column=phone_col).value = phones_str or None
                    if website_col and db_row.official_website:
                        ws.cell(row=r_idx, column=website_col).value = db_row.official_website
                    if conf_col and db_row.confidence is not None:
                        ws.cell(row=r_idx, column=conf_col).value = round(db_row.confidence, 3)
                    if updated_col:
                        ws.cell(row=r_idx, column=updated_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    if status_col:
                        if emails_str and phones_str:
                            ws.cell(row=r_idx, column=status_col).value = "Completed"
                        elif emails_str or phones_str:
                            ws.cell(row=r_idx, column=status_col).value = "Partially Enriched"
                        else:
                            ws.cell(row=r_idx, column=status_col).value = "No Contact Found"
                            
                    updated_count += 1
                    
            # 5. Save workbook back to both input folder and root directory (if they exist)
            wb.save(file_path)
            logger.info(f"[ExportManager] Successfully updated {updated_count} rows in original Excel file: {file_path}")
            
            root_path = "us_investors_enriched.xlsx"
            if os.path.exists(root_path) and os.path.abspath(root_path) != os.path.abspath(file_path):
                try:
                    wb.save(root_path)
                    logger.info(f"[ExportManager] Successfully mirrored {updated_count} rows to root Excel file: {root_path}")
                except Exception as e:
                    logger.warning(f"[ExportManager] Could not mirror updates to root Excel file: {e}")
            
        except Exception as e:
            logger.error(f"[ExportManager] In-place Excel update failed: {e}")
            raise e

